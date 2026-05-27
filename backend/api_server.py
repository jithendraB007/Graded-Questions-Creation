"""
api_server.py — FastAPI backend for the React frontend.

Start with:
    uvicorn backend.api_server:app --reload --port 8000
"""

import csv
import io
import json
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional, List

import openpyxl
import yaml
from openai import OpenAI
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

# Load .env
_env_file = ROOT / ".env"
if _env_file.exists():
    for _line in _env_file.read_text(encoding="utf-8").splitlines():
        _line = _line.strip()
        if _line and not _line.startswith("#") and "=" in _line:
            _k, _v = _line.split("=", 1)
            os.environ.setdefault(_k.strip(), _v.strip())

from generate_questions import build_prompt, load_samples, FOLDER_MAP, compute_bloom_targets
from backend.integrations.sheets import sheets_client

COURSES_DIR = ROOT / "courses"

# Difficulty → Bloom's level auto-mapping (K-taxonomy)
BLOOM_MAP: dict[str, str] = {
    "Easy":   "K2",   # Understand
    "Medium": "K3",   # Apply
    "Hard":   "K5",   # Evaluate
}


def _jaccard(a: str, b: str) -> float:
    """Word-level Jaccard similarity."""
    ta = set(re.findall(r'\w+', a.lower()))
    tb = set(re.findall(r'\w+', b.lower()))
    if not ta or not tb:
        return 0.0
    return len(ta & tb) / len(ta | tb)


def _has_4gram_overlap(a: str, b: str) -> bool:
    """True if any 4 consecutive words from a appear in the same order in b."""
    wa = re.findall(r'\w+', a.lower())
    wb = re.findall(r'\w+', b.lower())
    if len(wa) < 4 or len(wb) < 4:
        return False
    wb_4grams = {' '.join(wb[i:i+4]) for i in range(len(wb) - 3)}
    return any(' '.join(wa[i:i+4]) in wb_4grams for i in range(len(wa) - 3))


def _is_near_dup(text: str, pool: list[str], threshold: float = 0.65) -> bool:
    return any(
        _jaccard(text, existing) >= threshold or _has_4gram_overlap(text, existing)
        for existing in pool
    )

COURSE_DISPLAY = {
    "foundation":         "Communicative English Foundation",
    "advanced":           "Communicative English Advanced",
    "applied":            "Applied Communicative English",
    "language_analytics": "Language Analytics",
}


# ── Pydantic models ───────────────────────────────────────────────────────────

class GenerateRequest(BaseModel):
    course: str
    module: str
    topic: str
    question_type: str
    skill: str = ""
    count: int = 5
    difficulty: str = "Medium"
    marks: int = 2                     # marks per question — included in prompt
    bloom: str = ""                   # empty = auto-derive from difficulty via BLOOM_MAP
    course_outcome: str = ""
    model: str = "anthropic/claude-sonnet-4-5"
    existing_questions: List[str] = []  # pool question texts — used for dedup
    bloom_targets: List[str] = []     # per-question K-levels; computed if not provided


class Question(BaseModel):
    question: str
    solution: str
    explanation: str
    bloom: str = ""   # K1–K6, self-classified by the LLM in each response


class GenerateResponse(BaseModel):
    questions: list[Question]
    raw: str
    meta: dict
    filtered_count: int = 0


class DownloadQuestion(BaseModel):
    """Richer question record sent from the pool for Excel export."""
    question: str
    solution: str
    explanation: str
    bloom: str = ""
    difficulty: str = ""
    module_id: str = ""        # e.g. "module_2"
    module_display: str = ""
    topic_display: str = ""
    course_outcome: str = ""
    status: str = "pending"    # pending | approved | rejected
    feedback: str = ""         # reviewer's note on rejection / approval


class DownloadRequest(BaseModel):
    questions: list[DownloadQuestion]
    meta: dict                 # kept for filename generation only


class FeedbackItem(BaseModel):
    question: str
    solution: str
    explanation: str
    bloom: str = ""
    difficulty: str = ""
    question_type: str = ""
    course_outcome: str = ""
    status: str               # "approved" | "rejected"
    feedback: str = ""        # free-text reason on rejection


# ── Helpers ───────────────────────────────────────────────────────────────────

def _material_path(course: str, module: str, topic: str) -> Optional[Path]:
    mat_dir = COURSES_DIR / course / module / topic / "material"
    if not mat_dir.exists():
        return None
    files = sorted(mat_dir.glob("*.md"))
    return files[0] if files else None


def _load_meta(course: str, module: str, topic: str) -> dict:
    f = COURSES_DIR / course / module / topic / "metadata.json"
    if f.exists():
        try:
            return json.loads(f.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


def _module_display(course: str, module: str) -> str:
    mod_dir = COURSES_DIR / course / module
    for t in sorted(mod_dir.iterdir()):
        if t.is_dir():
            m = _load_meta(course, module, t.name)
            if m.get("module"):
                return m["module"]
    return f"Module {module.split('_')[-1]}"


def _topic_display(course: str, module: str, topic: str) -> str:
    m = _load_meta(course, module, topic)
    if m.get("topic_name"):
        return m["topic_name"]
    mat = _material_path(course, module, topic)
    if mat:
        return mat.stem.replace("_", " ").title()
    return topic.replace("_", " ").title()


def _parse_questions(raw: str, bloom_targets: list[str] | None = None) -> list[Question]:
    blocks = re.split(r"={3,}\s*QUESTION\s+\d+\s*={3,}", raw, flags=re.IGNORECASE)
    results = []
    q_index = 0
    for block in blocks:
        block = block.strip()
        if not block:
            continue

        def extract(label: str) -> str:
            m = re.search(
                rf"{label}:\s*\n(.*?)(?=\n(?:Question|Solution|Explanation|Bloom):|$)",
                block, re.DOTALL | re.IGNORECASE,
            )
            return m.group(1).strip() if m else ""

        def extract_bloom() -> str:
            m = re.search(r"Bloom(?:\s*Level)?:\s*(K[1-6])", block, re.IGNORECASE)
            return m.group(1).upper() if m else ""

        q = extract("Question")
        s = extract("Solution")
        e = extract("Explanation")
        b = extract_bloom()
        # Strip stray Bloom-level lines from explanation text
        e = re.sub(r'\n?Bloom(?:\s*Level)?:\s*K[1-6][^\n]*', '', e, flags=re.IGNORECASE).strip()
        e = re.sub(r'\bThis\s+(?:question\s+)?(?:tests?|demonstrates?|requires?|targets?)\s+K[1-6]\b[^\n]*', '', e, flags=re.IGNORECASE).strip()
        # Use per-question bloom target as fallback when LLM didn't self-classify
        if not b and bloom_targets and q_index < len(bloom_targets):
            b = bloom_targets[q_index]
        if q:
            results.append(Question(question=q, solution=s, explanation=e, bloom=b))
            q_index += 1
    return results


def _build_excel(questions: list[DownloadQuestion], meta: dict) -> bytes:
    wb = openpyxl.Workbook()

    # ── Sheet 1: Generated Questions (non-rejected) ───────────────────────
    ws = wb.active
    ws.title = "Generated Questions"

    headers = [
        "Q No.", "Module No.", "Bloom Level", "Difficulty",
        "Course Outcome", "Module Name", "Topic", "Question", "Solution", "Explanation",
    ]
    h_fill = PatternFill("solid", fgColor="1F4E79")
    h_font = Font(bold=True, color="FFFFFF", size=11)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = h_fill
        cell.font = h_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 30
    fill_a = PatternFill("solid", fgColor="DEEAF1")
    fill_b = PatternFill("solid", fgColor="FFFFFF")

    main_questions = [q for q in questions if q.status != "rejected"]
    for i, q in enumerate(main_questions, 1):
        mod_num = q.module_id.split("_")[-1] if q.module_id else meta.get("module", "").split("_")[-1]
        vals = [
            i,
            mod_num,
            q.bloom or meta.get("bloom", ""),
            q.difficulty or meta.get("difficulty", ""),
            q.course_outcome or meta.get("course_outcome", ""),
            q.module_display or meta.get("module_display", ""),
            q.topic_display or meta.get("topic_display", ""),
            q.question,
            q.solution,
            q.explanation,
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            cell.fill = fill_a if i % 2 == 0 else fill_b
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for col, w in enumerate([8, 10, 12, 12, 36, 28, 28, 60, 40, 50], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    ws.freeze_panes = "A2"

    # ── Sheet 2: Feedback (rejected + any question with a reviewer note) ─
    fb_questions = [q for q in questions if q.status == "rejected" or q.feedback.strip()]
    if fb_questions:
        wf = wb.create_sheet("Feedback")
        fb_headers = [
            "Q No.", "Module Name", "Topic", "Difficulty", "Bloom Level",
            "Question", "Solution", "Status", "Feedback Note",
        ]
        fb_h_fill = PatternFill("solid", fgColor="7B3F00")
        for col, h in enumerate(fb_headers, 1):
            cell = wf.cell(row=1, column=col, value=h)
            cell.fill = fb_h_fill
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        wf.row_dimensions[1].height = 30

        fill_rej = PatternFill("solid", fgColor="FFF2F2")
        fill_apv = PatternFill("solid", fgColor="F0FFF4")
        for i, q in enumerate(fb_questions, 1):
            vals = [
                i,
                q.module_display or meta.get("module_display", ""),
                q.topic_display or meta.get("topic_display", ""),
                q.difficulty or meta.get("difficulty", ""),
                q.bloom or meta.get("bloom", ""),
                q.question,
                q.solution,
                q.status,
                q.feedback,
            ]
            row_fill = fill_rej if q.status == "rejected" else fill_apv
            for col, val in enumerate(vals, 1):
                cell = wf.cell(row=i + 1, column=col, value=val)
                cell.fill = row_fill
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        for col, w in enumerate([8, 28, 28, 12, 12, 60, 40, 12, 50], 1):
            wf.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
        wf.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── FastAPI app ───────────────────────────────────────────────────────────────

app = FastAPI(title="GA Question Generator API", version="1.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["X-Drive-URL", "X-Drive-Error", "Content-Disposition"],
)


@app.get("/api/structure")
def get_structure():
    if not COURSES_DIR.exists():
        return {"courses": []}
    result = []
    for cd in sorted(COURSES_DIR.iterdir()):
        if not cd.is_dir():
            continue
        course = cd.name
        modules = []
        for md in sorted(cd.iterdir()):
            if not md.is_dir() or not md.name.startswith("module_"):
                continue
            module = md.name
            topics = []
            for td in sorted(md.iterdir()):
                if not td.is_dir() or not td.name.startswith("topic_"):
                    continue
                topic = td.name
                mat = _material_path(course, module, topic)
                meta = _load_meta(course, module, topic)
                topics.append({
                    "id": topic,
                    "display": _topic_display(course, module, topic),
                    "has_material": mat is not None,
                    "skills": meta.get("skills", []),
                })
            modules.append({
                "id": module,
                "display": _module_display(course, module),
                "topics": topics,
            })
        result.append({
            "id": course,
            "display": COURSE_DISPLAY.get(course, course.replace("_", " ").title()),
            "modules": modules,
        })
    return {"courses": result}


@app.post("/api/generate", response_model=GenerateResponse)
def generate(req: GenerateRequest):
    mat = _material_path(req.course, req.module, req.topic)
    if not mat:
        raise HTTPException(404, "No reading material found for this topic.")

    api_key = os.environ.get("OPENROUTER_API_KEY", "")
    if not api_key or api_key.startswith("sk-or-paste"):
        raise HTTPException(
            400,
            "OpenRouter API key not configured. Open D:\\GA\\.env, set OPENROUTER_API_KEY, then restart the backend."
        )

    # Auto-derive Bloom level from difficulty if not explicitly provided
    bloom = req.bloom if req.bloom else BLOOM_MAP.get(req.difficulty, "K3")

    # Compute per-question bloom targets (use client-provided list if given)
    targets = (
        req.bloom_targets
        if req.bloom_targets and len(req.bloom_targets) == req.count
        else compute_bloom_targets(req.course_outcome, req.difficulty, req.count)
    )

    material_text = mat.read_text(encoding="utf-8")
    material_text = re.sub(r"<[^>]+>", " ", material_text)
    material_text = re.sub(r"\n{3,}", "\n\n", material_text).strip()

    samples = load_samples(req.question_type)
    prompt = build_prompt(
        material=material_text,
        question_type=req.question_type,
        count=req.count,
        bloom=bloom,
        difficulty=req.difficulty,
        course_outcome=req.course_outcome,
        samples=samples,
        existing_questions=req.existing_questions or None,
        bloom_targets=targets,
        marks=req.marks,
    )

    client = OpenAI(
        base_url="https://openrouter.ai/api/v1",
        api_key=api_key,
    )
    try:
        resp = client.chat.completions.create(
            model=req.model,
            max_tokens=8192,
            messages=[{"role": "user", "content": prompt}],
        )
    except Exception as e:
        err = str(e)
        if "insufficient" in err.lower() or "balance" in err.lower() or "quota" in err.lower():
            raise HTTPException(402, "OpenRouter credit balance too low. Top up at openrouter.ai/credits")
        if "invalid" in err.lower() or "authentication" in err.lower() or "401" in err:
            raise HTTPException(401, "Invalid OpenRouter API key. Check D:\\GA\\.env")
        raise HTTPException(500, f"Generation failed: {err}")

    raw = resp.choices[0].message.content
    questions = _parse_questions(raw, bloom_targets=targets)

    # ── Dedup: remove near-duplicates within the new batch ───────────────────
    # Use Jaccard-only at 0.85 (NOT 4-gram) — 4-gram triggers false positives
    # when all questions share the same instruction template (e.g. "Rearrange
    # the following sentences" or "Fill in the blanks").
    if len(questions) > 1:
        unique: list[Question] = [questions[0]]
        for q in questions[1:]:
            if not any(_jaccard(q.question, u.question) >= 0.85 for u in unique):
                unique.append(q)
        questions = unique

    # ── Dedup: remove near-duplicates against existing pool ──────────────────
    # Use Jaccard >= 0.85 only (same as within-batch). 4-gram overlap causes
    # false positives on instruction-based types (e.g. "Rearrange the following
    # sentences" appears in every Sentence Arrangement question).
    filtered_count = 0
    if req.existing_questions:
        seen = list(req.existing_questions)
        final: list[Question] = []
        for q in questions:
            if any(_jaccard(q.question, e) >= 0.85 for e in seen):
                filtered_count += 1
            else:
                final.append(q)
                seen.append(q.question)
        questions = final

    meta = {
        "course": req.course,
        "course_display": COURSE_DISPLAY.get(req.course, req.course),
        "module": req.module,
        "module_display": _module_display(req.course, req.module),
        "topic": req.topic,
        "topic_display": _topic_display(req.course, req.module, req.topic),
        "question_type": req.question_type,
        "bloom": bloom,
        "difficulty": req.difficulty,
        "course_outcome": req.course_outcome,
        "material_file": mat.name,
    }
    return GenerateResponse(questions=questions, raw=raw, meta=meta, filtered_count=filtered_count)


@app.post("/api/download/excel")
def download_excel(req: DownloadRequest):
    data = _build_excel(req.questions, req.meta)

    # Build a descriptive filename: CourseName_Module_N_QuestionType_YYYY-MM-DD.xlsx
    course_word = (req.meta.get("course_display") or req.meta.get("course", "Course")).split()[0]
    mod_num     = req.meta.get("module", "module_1").split("_")[-1]
    qtype       = req.meta.get("question_type", "Questions").replace(" ", "_").replace("/", "-")
    date_str    = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    fname       = f"{course_word}_Module_{mod_num}_{qtype}_{date_str}.xlsx"

    # Upload to Google Drive if authenticated (download always proceeds regardless)
    drive_url, drive_err = "", ""
    if sheets_client.auth_status == "ready":
        drive_url, drive_err = sheets_client.upload_excel_to_drive(data, fname)

    resp_headers = {"Content-Disposition": f"attachment; filename={fname}"}
    if drive_url:
        resp_headers["X-Drive-URL"] = drive_url
    if drive_err:
        resp_headers["X-Drive-Error"] = drive_err[:200]

    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=resp_headers,
    )


@app.get("/api/samples")
def get_samples():
    """Return sample question counts per question type."""
    result = {}
    for qtype, folder in FOLDER_MAP.items():
        eval_file = ROOT / "evals" / folder / "eval_tests.yaml"
        count = 0
        if eval_file.exists():
            try:
                data = yaml.safe_load(eval_file.read_text(encoding="utf-8")) or []
                count = len(data)
            except Exception:
                count = 0
        result[qtype] = count
    return result


@app.post("/api/samples/upload")
async def upload_samples(
    question_type: str = Form(...),
    file: UploadFile = File(...),
):
    """Append rows from a CSV upload to the question type's eval_tests.yaml."""
    folder = FOLDER_MAP.get(question_type)
    if not folder:
        raise HTTPException(400, f"Unknown question type: {question_type!r}")

    content = (await file.read()).decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(content))

    eval_file = ROOT / "evals" / folder / "eval_tests.yaml"
    eval_file.parent.mkdir(parents=True, exist_ok=True)

    existing: list = []
    if eval_file.exists():
        try:
            existing = yaml.safe_load(eval_file.read_text(encoding="utf-8")) or []
        except Exception:
            existing = []

    added = 0
    for row in reader:
        question    = (row.get("Question") or row.get("question") or "").strip()
        solution    = (row.get("Solution") or row.get("solution") or "").strip()
        explanation = (row.get("Explanation") or row.get("explanation") or "").strip()
        bloom       = (row.get("Bloom's Level") or row.get("Bloom Level") or row.get("bloom") or "").strip()
        difficulty  = (row.get("Difficulty") or row.get("difficulty") or "").strip()
        co          = (row.get("CO") or row.get("Course Outcome") or "").strip()
        if not question:
            continue
        existing.append({
            "vars": {
                "question_type": question_type,
                "bloom_level":   bloom,
                "difficulty":    difficulty,
                "course_outcome": co,
                "question":      question,
                "solution":      solution,
                "explanation":   explanation,
            },
            "assert": [{
                "type":  "llm-rubric",
                "value": f"Evaluate this {question_type} question for correctness and clarity. Respond: VALID or INVALID.",
            }],
        })
        added += 1

    eval_file.write_text(
        yaml.dump(existing, allow_unicode=True, default_flow_style=False),
        encoding="utf-8",
    )
    return {"added": added, "total": len(existing), "question_type": question_type}


# ── Module-level batch generation ────────────────────────────────────────────

class GenerateModuleRequest(BaseModel):
    course: str
    module: str
    question_type: str
    count: int = 3             # questions per topic
    difficulty: str = "Medium"
    course_outcome: str = ""
    model: str = "anthropic/claude-sonnet-4-5"
    existing_questions: List[str] = []


@app.post("/api/generate/module")
def generate_module(req: GenerateModuleRequest):
    """Generate questions for every topic in a module that has material."""
    api_key = os.environ.get("OPENROUTER_API_KEY", "")
    if not api_key or api_key.startswith("sk-or-paste"):
        raise HTTPException(400, "OpenRouter API key not configured.")

    mod_dir = COURSES_DIR / req.course / req.module
    if not mod_dir.exists():
        raise HTTPException(404, f"Module not found: {req.module}")

    bloom = BLOOM_MAP.get(req.difficulty, "K3")
    client = OpenAI(base_url="https://openrouter.ai/api/v1", api_key=api_key)
    samples = load_samples(req.question_type)

    all_questions: list[dict] = []
    seen_texts: list[str] = list(req.existing_questions)
    topic_results: list[dict] = []

    for td in sorted(mod_dir.iterdir()):
        if not td.is_dir() or not td.name.startswith("topic_"):
            continue
        topic = td.name
        mat = _material_path(req.course, req.module, topic)
        if not mat:
            continue

        targets = compute_bloom_targets(req.course_outcome, req.difficulty, req.count)
        material_text = mat.read_text(encoding="utf-8")
        material_text = re.sub(r"<[^>]+>", " ", material_text)
        material_text = re.sub(r"\n{3,}", "\n\n", material_text).strip()

        prompt = build_prompt(
            material=material_text,
            question_type=req.question_type,
            count=req.count,
            bloom=bloom,
            difficulty=req.difficulty,
            course_outcome=req.course_outcome,
            samples=samples,
            existing_questions=seen_texts or None,
            bloom_targets=targets,
        )

        try:
            resp = client.chat.completions.create(
                model=req.model,
                max_tokens=8192,
                messages=[{"role": "user", "content": prompt}],
            )
        except Exception as e:
            topic_results.append({"topic": topic, "error": str(e), "questions": []})
            continue

        raw = resp.choices[0].message.content
        questions = _parse_questions(raw, bloom_targets=targets)

        # Dedup within batch and against seen pool
        unique: list[Question] = []
        for q in questions:
            if not _is_near_dup(q.question, seen_texts):
                unique.append(q)
                seen_texts.append(q.question)

        topic_display = _topic_display(req.course, req.module, topic)
        module_display = _module_display(req.course, req.module)
        meta = {
            "course": req.course,
            "module": req.module,
            "module_display": module_display,
            "topic": topic,
            "topic_display": topic_display,
            "question_type": req.question_type,
            "bloom": bloom,
            "difficulty": req.difficulty,
            "course_outcome": req.course_outcome,
        }
        for q in unique:
            all_questions.append({**q.model_dump(), **meta})

        topic_results.append({
            "topic": topic,
            "topic_display": topic_display,
            "count": len(unique),
            "questions": [q.model_dump() for q in unique],
            "meta": meta,
        })

    return {"topics": topic_results, "total": len(all_questions)}


# ── Feedback & DSPy optimization ─────────────────────────────────────────────

FEEDBACK_DIR = ROOT / "feedback"


@app.post("/api/feedback")
def submit_feedback(items: List[FeedbackItem]):
    """Persist approved/rejected questions as training signal."""
    FEEDBACK_DIR.mkdir(parents=True, exist_ok=True)
    feedback_file = FEEDBACK_DIR / "feedback.jsonl"
    with feedback_file.open("a", encoding="utf-8") as f:
        for item in items:
            f.write(json.dumps(item.model_dump(), ensure_ascii=False) + "\n")
    return {"stored": len(items)}


@app.post("/api/optimize")
def optimize_prompts():
    """Bootstrap few-shot examples from approved feedback using DSPy-style selection.

    Approved questions are added to eval_tests.yaml for each question type,
    improving future few-shot prompts. Rejected feedback is logged for review.
    Returns a summary of what changed.
    """
    feedback_file = FEEDBACK_DIR / "feedback.jsonl"
    if not feedback_file.exists():
        raise HTTPException(404, "No feedback collected yet. Approve/reject some questions first.")

    approved: list[dict] = []
    rejected: list[dict] = []
    with feedback_file.open(encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                item = json.loads(line)
                (approved if item.get("status") == "approved" else rejected).append(item)
            except Exception:
                pass

    if not approved:
        return {"message": "No approved examples yet.", "rejected_count": len(rejected)}

    # Group by question_type and select up to 5 best examples per type
    by_type: dict[str, list[dict]] = {}
    for a in approved:
        qt = a.get("question_type", "")
        by_type.setdefault(qt, []).append(a)

    updated_types: list[str] = []
    for qt, items in by_type.items():
        folder = FOLDER_MAP.get(qt)
        if not folder:
            continue
        eval_file = ROOT / "evals" / folder / "eval_tests.yaml"
        eval_file.parent.mkdir(parents=True, exist_ok=True)
        existing: list = []
        if eval_file.exists():
            existing = yaml.safe_load(eval_file.read_text(encoding="utf-8")) or []

        added = 0
        existing_texts = {e.get("vars", {}).get("question", "") for e in existing}
        for item in items[:5]:
            if item.get("question", "") in existing_texts:
                continue
            existing.append({
                "vars": {
                    "question_type": qt,
                    "bloom_level":   item.get("bloom", ""),
                    "difficulty":    item.get("difficulty", ""),
                    "course_outcome": item.get("course_outcome", ""),
                    "question":      item.get("question", ""),
                    "solution":      item.get("solution", ""),
                    "explanation":   item.get("explanation", ""),
                },
                "assert": [{"type": "llm-rubric",
                            "value": f"Evaluate this {qt} question for correctness and clarity. Respond: VALID or INVALID."}],
            })
            added += 1
        if added:
            eval_file.write_text(yaml.dump(existing, allow_unicode=True, default_flow_style=False), encoding="utf-8")
            updated_types.append(qt)

    # Archive used feedback so it isn't re-applied on next optimize
    archive = FEEDBACK_DIR / "feedback_archived.jsonl"
    with archive.open("a", encoding="utf-8") as out, feedback_file.open(encoding="utf-8") as inp:
        out.write(inp.read())
    feedback_file.unlink()

    return {
        "approved_added": sum(min(5, len(v)) for v in by_type.values()),
        "rejected_logged": len(rejected),
        "types_updated": updated_types,
        "message": f"Prompt examples updated for: {', '.join(updated_types) or 'none'}",
    }


# ── Google Sheets integration ─────────────────────────────────────────────────

class SheetsLogRequest(BaseModel):
    questions: list[dict]          # DownloadQuestion-shaped dicts from the pool


@app.get("/api/sheets/status")
def sheets_status():
    """Return auth status and spreadsheet info."""
    status = sheets_client.auth_status
    info   = sheets_client.get_spreadsheet_info() if status == "ready" else {}
    return {"auth_status": status, "auth_error": sheets_client._auth_error, **info}


@app.post("/api/sheets/auth")
def sheets_auth():
    """Trigger Google OAuth flow (opens browser on the server machine)."""
    return sheets_client.start_auth()


@app.post("/api/sheets/log")
def sheets_log(req: SheetsLogRequest):
    """Append questions to Google Sheets and refresh Dashboard."""
    if sheets_client.auth_status != "ready":
        raise HTTPException(401, "Not authenticated with Google Sheets. Call /api/sheets/auth first.")
    try:
        result = sheets_client.log_questions(req.questions)
        return result
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/api/sheets/dashboard")
def sheets_dashboard():
    """Return aggregated stats from the Google Sheets Questions Log."""
    if sheets_client.auth_status != "ready":
        raise HTTPException(401, "Not authenticated with Google Sheets.")
    try:
        return sheets_client.get_stats()
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/api/sheets/debug")
def sheets_debug():
    """Return detailed credential info to help diagnose Sheets auth problems."""
    from backend.integrations.sheets import TOKEN_FILE, CONFIG_FILE
    creds_info: dict = {}
    try:
        creds = sheets_client._load_creds()
        if creds:
            creds_info = {
                "valid":             creds.valid,
                "expired":           creds.expired,
                "has_refresh_token": bool(creds.refresh_token),
                "expiry":            str(creds.expiry) if creds.expiry else None,
                "scopes":            list(creds.scopes or []),
            }
        else:
            creds_info = {"loaded": False}
    except Exception as e:
        creds_info = {"error": str(e)}
    return {
        "auth_status":        sheets_client.auth_status,
        "auth_error":         sheets_client._auth_error,
        "token_file_exists":  TOKEN_FILE.exists(),
        "config_file_exists": CONFIG_FILE.exists(),
        "google_token_env":   bool(os.environ.get("GOOGLE_TOKEN")),
        "spreadsheet_id_env": bool(os.environ.get("GOOGLE_SPREADSHEET_ID")),
        "credentials":        creds_info,
    }


@app.post("/api/sheets/bulk-approve")
def sheets_bulk_approve():
    """Mark every pending question in the Questions Log as approved."""
    if sheets_client.auth_status != "ready":
        raise HTTPException(401, "Not authenticated with Google Sheets.")
    try:
        result = sheets_client.bulk_approve()
        return result
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/api/sheets/token")
def sheets_token():
    """Return the current refreshed token JSON — use this to update GOOGLE_TOKEN in Render."""
    creds = sheets_client._load_creds()
    if not creds:
        raise HTTPException(401, "Not authenticated — no valid token available.")
    return {"token_json": creds.to_json(), "expiry": str(creds.expiry) if creds.expiry else None}


@app.post("/api/import/csv")
async def import_csv(file: UploadFile = File(...), status: str = "approved"):
    """Parse a question CSV and append all rows to Google Sheets Questions Log."""
    if sheets_client.auth_status != "ready":
        raise HTTPException(401, "Not authenticated with Google Sheets. Sign in first.")

    content = await file.read()
    # Handle UTF-8 BOM that Excel adds
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))

    def _detect_type(q_text: str) -> str:
        ql = q_text.lower()
        if any(k in ql for k in ["error correction", "identify and correct",
                                   "correct the error", "find and correct",
                                   "underline the error", "rewrite the corrected",
                                   "contains one error"]):
            return "Error Correction"
        return "Fill in the Blanks"

    items = []
    for row in reader:
        q_text = row.get("Questions", "").strip()
        if not q_text:
            continue
        items.append({
            "question":      q_text,
            "solution":      row.get("Solution", "").strip(),
            "explanation":   "",
            "bloom":         row.get("Blooms Level", "").strip(),
            "difficulty":    row.get("Difficulty level", "").strip(),
            "question_type": _detect_type(q_text),
            "module_id":     row.get("Module Number", "").strip(),
            "module_display": row.get("Name of the Module", "").strip(),
            "topic_display": row.get("Topic name", "").strip(),
            "course_outcome": row.get("Course Outcomes", "").strip(),
            "course_display": "",
            "status":        status if status in ("approved", "pending", "rejected") else "approved",
            "feedback":      "",
        })

    if not items:
        raise HTTPException(400, "No question rows found in the CSV.")

    result = sheets_client.log_questions(items)
    return {"logged": result.get("logged", len(items)), "total": len(items)}


@app.get("/health")
def health():
    api_key = os.environ.get("OPENROUTER_API_KEY", "")
    return {
        "status": "ok",
        "api_key_configured": bool(api_key and not api_key.startswith("sk-or-paste")),
    }


# ── Syllabus endpoints ────────────────────────────────────────────────────────

SYLLABI_DIR = ROOT / "syllabi"


def _extract_text_from_upload(file_bytes: bytes, filename: str) -> str:
    ext = (filename.rsplit(".", 1)[-1] if "." in filename else "txt").lower()
    if ext == "pdf":
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(file_bytes))
        return "\n".join(page.extract_text() or "" for page in reader.pages)
    elif ext in ("docx", "doc"):
        from docx import Document as DocxDoc
        doc = DocxDoc(io.BytesIO(file_bytes))
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    else:
        return file_bytes.decode("utf-8-sig", errors="replace")


_SYLLABUS_PROMPT = """You are an expert curriculum analyst. Extract the course structure from the syllabus text below and return it as a single JSON object matching this exact schema:

{{
  "course_name": "<full course name>",
  "co_definitions": {{
    "CO1": "<full outcome text>",
    "CO2": "<full outcome text>"
  }},
  "units": [
    {{
      "unit_number": 1,
      "unit_name": "<unit / module name>",
      "co": "CO1",
      "topics": [
        {{
          "topic_name": "<topic name>",
          "recommended_question_types": ["Fill in the Blanks", "Error Correction"],
          "marks": 2,
          "description": "<brief description>"
        }}
      ]
    }}
  ],
  "notes": "<any extra information>"
}}

Rules:
- Map each unit to the CO that best covers it (CO1 for unit 1, CO2 for unit 2, etc. unless the syllabus says otherwise).
- Recommended question types must only come from this list:
  Fill in the Blanks, Cloze, Error Correction, Sentence Arrangement,
  Jumbled Sentences, Jumbled Words, Sentence Conversion, Sentence Correction / MCQ
- If the syllabus does not specify question types for a topic, infer reasonable ones based on the topic content.
- Return ONLY the JSON object — no markdown fences, no explanation text.

SYLLABUS TEXT:
{syllabus_text}"""


@app.post("/api/syllabus/upload")
async def upload_syllabus(
    course_id: str = Form(...),
    file: UploadFile = File(...),
):
    """Upload a syllabus PDF/DOCX/TXT and extract course structure via LLM."""
    api_key = os.environ.get("OPENROUTER_API_KEY", "")
    if not api_key or api_key.startswith("sk-or-paste"):
        raise HTTPException(400, "OpenRouter API key not configured.")

    file_bytes = await file.read()
    try:
        text = _extract_text_from_upload(file_bytes, file.filename or "syllabus.txt")
    except Exception as e:
        raise HTTPException(422, f"Could not read file: {e}")

    if not text.strip():
        raise HTTPException(422, "No text could be extracted from the file.")

    prompt = _SYLLABUS_PROMPT.format(syllabus_text=text[:12000])

    client = OpenAI(base_url="https://openrouter.ai/api/v1", api_key=api_key)
    try:
        resp = client.chat.completions.create(
            model="anthropic/claude-sonnet-4-5",
            max_tokens=8192,
            messages=[{"role": "user", "content": prompt}],
        )
    except Exception as e:
        raise HTTPException(500, f"LLM extraction failed: {e}")

    raw_json = resp.choices[0].message.content.strip()
    # Strip markdown fences if model added them
    raw_json = re.sub(r"^```(?:json)?\s*", "", raw_json)
    raw_json = re.sub(r"\s*```$", "", raw_json)

    try:
        structure = json.loads(raw_json)
    except json.JSONDecodeError as e:
        raise HTTPException(500, f"LLM returned invalid JSON: {e}\n\nRaw:\n{raw_json[:500]}")

    SYLLABI_DIR.mkdir(parents=True, exist_ok=True)
    out_file = SYLLABI_DIR / f"{course_id}.json"
    out_file.write_text(json.dumps(structure, ensure_ascii=False, indent=2), encoding="utf-8")

    return {"course_id": course_id, "structure": structure}


@app.get("/api/syllabus/{course_id}")
def get_syllabus(course_id: str):
    """Return the parsed syllabus structure for a course."""
    f = SYLLABI_DIR / f"{course_id}.json"
    if not f.exists():
        raise HTTPException(404, f"No syllabus found for course '{course_id}'")
    try:
        return json.loads(f.read_text(encoding="utf-8"))
    except Exception as e:
        raise HTTPException(500, f"Could not read syllabus: {e}")


@app.get("/api/syllabi")
def list_syllabi():
    """List all courses that have an uploaded syllabus."""
    if not SYLLABI_DIR.exists():
        return []
    result = []
    for f in sorted(SYLLABI_DIR.glob("*.json")):
        try:
            data = json.loads(f.read_text(encoding="utf-8"))
            result.append({
                "course_id": f.stem,
                "course_name": data.get("course_name", f.stem),
                "unit_count": len(data.get("units", [])),
                "co_count": len(data.get("co_definitions", {})),
            })
        except Exception:
            result.append({"course_id": f.stem, "course_name": f.stem})
    return result


# ── Serve built React frontend (production) ───────────────────────────────────
_frontend_dist = ROOT / "frontend" / "dist"
if _frontend_dist.exists():
    app.mount("/", StaticFiles(directory=str(_frontend_dist), html=True), name="frontend")
